"""
Unified Roster Scheduler
========================
Config-driven CP-SAT scheduler: MO, HO, REG roster types.
Roster-specific logic in subclasses only.

Usage:
    from roster_scheduler import build_scheduler, ROSTER_CONFIGS
    import copy

    cfg = copy.deepcopy(ROSTER_CONFIGS["MO"])
    cfg["fairness_pools"][0]["metrics"]["pts"] = 30

    s = build_scheduler("MO", "input.csv", "outputs/", cfg)
    s.set_public_holidays(["1-May-26"])
    s.load_data()
    df = s.solve_and_export()
"""

import os, copy, shutil
from datetime import datetime, timedelta

import pandas as pd
from ortools.sat.python import cp_model
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

_MONTH_MAP = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
              "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}



# ─────────────────────────────────────────────────────────────────────────────
# DEFAULT CONFIGS
# ─────────────────────────────────────────────────────────────────────────────

# shift_categories keys and their scheduler behaviour:
#   main          — each named shift filled exactly once every day
#   weekend_extra — each named shift filled exactly once on weekends/PH only
#   cicu          — filled exactly once every day; gated via EligibleShifts
#   hybrid_shift  — filled exactly once every day; weekday half-call, weekend full call (when names non-empty)
#   wr            — named slots filled in order up to slots[day_type];
#                   ph_after_sunday overrides slots["ph"] when PH follows Sunday
#   sb            — total of named shifts == slots[day_type]
#   half          — total of named shifts == slots[day_type]

ROSTER_CONFIGS = {
    "MO": {
        "roster_type": "MO",
        "label": "Medical Officer Roster",
        "shift_categories": {
            "main":          {"names": ["MO1","MO2","MO3","MO4"]},
            "weekend_extra": {"names": ["MO5"]},
            "cicu":          {"names": ["MO6"]},
            "hybrid_shift": {"names": []},
            "wr":  {"names": ["WR1","WR2","WR3"],
                    "slots": {"weekday":0,"saturday":3,"sunday":2,"ph":3},
                    "ph_after_sunday": 2},
            "sb":  {"names": ["SB"],
                    "slots": {"weekday":1,"saturday":1,"sunday":1,"ph":1}},
            "half":{"names": [],
                    "slots": {"weekday":0,"saturday":0,"sunday":0,"ph":0}},
        },
        "call_points": {
            "saturday":4,"ph":4,"sunday":3,"friday":3,"pre_ph":3,"weekday":1,
        },
        "call_points_scale": 1,
        "soft_penalties": {"BLOCK":30,"REQUEST":30},
        "hard_blocks":    ["POSTCALL","COURSE","SUBSPEC","CLINIC","AUTOBLOCK"],
        "leave_values":   ["LEAVE"],
        "limits":         {"wr_max":2,"sb_max":2},
        "eligible_shifts_mode": "cicushift_only",
        "new_block_days": 10,
        "new_phantom_points": 4,
        "vikas_days":     ["thursday","sunday"],
        "team_preferences": {
            "A":["MO1"],"HA":["MO1"],"B":["MO2"],
            "C":["MO3"],"D":["MO3","MO4"],"CICU":["MO6"],
        },
        "team_pref_penalty": 2,
        "fairness_pools": [{
            "label": "Regular Staff",
            "exclude_subtypes": ["FAMMED","ELECTIVES"],
            "exclude_tags":     ["HC","VIKAS"],
            "metrics": {
                "pts":20,"full_calls":6,"sat_calls":3,
                "wr_count":2,"sb_count":2,
                "golden_wknds":0,"shift_spacing":0,"cross_month":0,
            },
        }],
        "required_cols": ["Name","Team","Ward","Subspec","EligibleShifts","SpecialReq"],
        "date_col_start": 6,
        "export_skip_cols": ["SpecialReq"],
        "color_priority": ["LEAVE","POSTCALL","COURSE","SUBSPEC","AUTOBLOCK","REQUEST","BLOCK"],
    },

    "HO": {
        "roster_type": "HO",
        "label": "House Officer Roster",
        "shift_categories": {
            "main":          {"names": ["HO1","HO2","HO3","HO4","HO5"]},
            "weekend_extra": {"names": []},
            "cicu":          {"names": []},
            "hybrid_shift": {"names": []},
            "wr":  {"names": ["WR1","WR2","WR3"],
                    "slots": {"weekday":0,"saturday":3,"sunday":3,"ph":3},
                    "ph_after_sunday": 3},
            "sb":  {"names": ["SB"],
                    "slots": {"weekday":1,"saturday":1,"sunday":1,"ph":1}},
            "half":{"names": ["HO6"],
                    "slots": {"weekday":0,"saturday":1,"sunday":1,"ph":1}},
        },
        "call_points": {
            "saturday":5,"ph":5,"sunday":4,"friday":3,"pre_ph":3,"weekday":2,
        },
        "call_points_scale": 2,
        "ho6_points": 2,   # flat 1.0pt stored ×2
        "soft_penalties": {"BLOCK":50,"REQUEST":50,"NEW":10},
        "hard_blocks":    ["POSTCALL","LEAVE","AUTOBLOCK"],
        "leave_values":   ["LEAVE"],
        "limits":         {"wr_max":2,"sb_max":2},
        "eligible_shifts_mode": "none",
        "new_block_days": 0,
        "new_phantom_points": 0,
        "vikas_days":     [],
        "team_preferences": {
            "A":["HO1"],"HA":["HO1"],"B":["HO2"],"C":["HO3"],"D":["HO3","HO4"],
        },
        "team_pref_penalty": 2,
        "fairness_pools": [{
            "label": "All HO Staff",
            "exclude_subtypes": [],
            "exclude_tags":     [],
            "metrics": {
                "pts":30,"cross_month":15,"golden_wknds":5,"shift_spacing":5,
                "wr_count":1,"sb_count":1,"full_calls":10,"sat_calls":0,
            },
        }],
        "required_cols": ["Name","Team","Ward","SpecialReq"],
        "date_col_start": 4,
        "color_priority": ["LEAVE","POSTCALL","AUTOBLOCK","REQUEST","BLOCK","NEW"],
    },

    "REG": {
        "roster_type": "REG",
        "label": "Registrar Roster",
        "shift_categories": {
            "main":          {"names": ["R1","R2"]},
            "weekend_extra": {"names": []},
            "cicu":          {"names": ["R4"]},
            "hybrid_shift": {"names": ["R3"]},
            "wr":  {"names": ["WR1","WR2"],
                    "slots": {"weekday":0,"saturday":2,"sunday":2,"ph":2},
                    "ph_after_sunday": 2},
            "sb":  {"names": [],
                    "slots": {"weekday":0,"saturday":0,"sunday":0,"ph":0}},
            "half":{"names": [],
                    "slots": {"weekday":0,"saturday":0,"sunday":0,"ph":0}},
        },
        "call_points": {
            "saturday":8,"ph":6,"sunday":6,"friday":4,"weekday":2,
        },
        "r3_points": {
            "saturday":8,"ph":6,"sunday":6,"friday":2,"weekday":1,
        },
        "call_points_scale": 2,
        "soft_penalties": {"BLOCK":30,"REQUEST":30},
        "hard_blocks":    ["POSTCALL","COURSE","SUBSPEC","CLINIC","AUTOBLOCK"],
        "leave_values":   ["LEAVE"],
        "limits":         {"wr_max":2,"sb_max":0,"rp_r3_max":3,"ac_call_min":1,"ac_call_max":2},
        "eligible_shifts_mode": "whitelist",
        "new_block_days": 0,
        "new_phantom_points": 0,
        "vikas_days":     [],
        "team_preferences": {},
        "team_pref_penalty": 2,
        "fairness_pools": [
            {
                "label": "SR Pool",
                "include_types":    ["SR"],
                "exclude_subtypes": [],
                "exclude_tags":     [],
                "metrics": {"pts":20,"sat_calls":3,"wr_count":2,"full_calls":0,"golden_wknds":0},
            },
            {
                "label": "RP Pool",
                "include_types":    ["RP"],
                "exclude_subtypes": [],
                "exclude_tags":     [],
                "metrics": {"pts":20,"wr_count":2,"sat_calls":0},
            },
        ],
        "required_cols": ["Name","StaffType","Subspec","EligibleShifts","SpecialReq"],
        "date_col_start": 5,
        "color_priority": ["LEAVE","POSTCALL","COURSE","SUBSPEC","AUTOBLOCK","REQUEST","BLOCK"],
    },
}


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _parse_date(dc: str) -> datetime | None:
    """Parse date column header → datetime. Supports '6-Jan', '6 Jan', DD/MM/YYYY."""
    dc = str(dc).strip()
    try:
        for sep in ("-", " "):
            parts = dc.split(sep)
            if len(parts) == 2 and parts[1] in _MONTH_MAP:
                m = _MONTH_MAP[parts[1]]
                today = datetime.today()
                year  = today.year if m >= today.month else today.year + 1
                return datetime(year, m, int(parts[0]))
        if dc.count("/") == 2:
            d, m, y = dc.split("/")
            y = int(y); y += 2000 if y < 100 else 0
            return datetime(y, int(m), int(d))
        r = pd.to_datetime(dc, errors="coerce")
        return None if pd.isna(r) else r.to_pydatetime()
    except Exception:
        return None


def _ph_key(dt: datetime) -> str:
    return f"{dt.day}-{dt.strftime('%b')}-{str(dt.year)[-2:]}"


# Hardcoded Singapore public holidays — update annually from MOM gazette.
_SG_PUBLIC_HOLIDAYS: dict[int, list[str]] = {
    2025: [
        "1-Jan-25","29-Jan-25","30-Jan-25","31-Mar-25","18-Apr-25",
        "1-May-25","12-May-25","7-Jun-25","9-Aug-25","20-Oct-25",
        "25-Oct-25","25-Dec-25",
    ],
    2026: [
        "1-Jan-26","17-Feb-26","18-Feb-26","3-Apr-26","1-May-26",
        "19-May-26","26-May-26","9-Aug-26","10-Nov-26","25-Dec-26",
    ],
}


def get_singapore_ph(year: int) -> list[str]:
    """Return hardcoded SG public holidays for year."""
    if year not in _SG_PUBLIC_HOLIDAYS:
        raise ValueError(
            f"No public holiday data for {year}. "
            f"Add it to _SG_PUBLIC_HOLIDAYS in roster_scheduler.py."
        )
    return list(_SG_PUBLIC_HOLIDAYS[year])


def fetch_singapore_ph(year: int) -> list[str]:
    """Optional live fetch from data.gov.sg (collection 691).
    Returns fetched list on success, [] on any failure.
    Use get_singapore_ph() as the primary source.
    """
    _COLLECTION = "691"
    _META_URL   = f"https://api-production.data.gov.sg/v2/public/api/collections/{_COLLECTION}/metadata"
    _SEARCH_URL = "https://data.gov.sg/api/action/datastore_search?resource_id={rid}&limit=20"

    try:
        import json, ssl, time, urllib.request
        try:
            import certifi
            ctx = ssl.create_default_context(cafile=certifi.where())
        except ImportError:
            ctx = ssl.create_default_context()

        def _get(url):
            req = urllib.request.Request(url, headers={"User-Agent": "RosterApp/1.0"})
            with urllib.request.urlopen(req, timeout=8, context=ctx) as r:
                return json.loads(r.read())

        meta     = _get(_META_URL)
        children = meta["data"]["collectionMetadata"]["childDatasets"]

        for rid in reversed(children):
            try:
                data    = _get(_SEARCH_URL.format(rid=rid))
                records = data["result"]["records"]
                if not records:
                    continue
                if datetime.strptime(records[0]["date"], "%Y-%m-%d").year != year:
                    continue
                result = [_ph_key(datetime.strptime(r["date"], "%Y-%m-%d")) for r in records]
                print(f"[PH fetch] {year}: {len(result)} holidays from dataset {rid}")
                return result
            except Exception:
                time.sleep(0.3)
                continue

        print(f"[PH fetch] No dataset found for {year}")
        return []

    except Exception as e:
        print(f"[PH fetch] Failed: {e}")
        return []


def parse_phantom_points(special_req, phantom_range=(1,10)) -> tuple[list, int]:
    """Parse SpecialReq → (tags, phantom_int).
    MINUS{n} → +n phantom (fewer calls). PLUS{n} → -n phantom (more calls).
    """
    if not special_req or pd.isna(special_req):
        return [], 0
    lo, hi = phantom_range
    tags, phantom = [], 0
    for p in str(special_req).upper().split(","):
        p = p.strip()
        for prefix, sign in (("MINUS", 1), ("PLUS", -1)):
            if p.startswith(prefix) and len(p) > len(prefix):
                try:
                    n = int(p[len(prefix):])
                    if lo <= n <= hi:
                        phantom += sign * n
                except ValueError:
                    pass
                break
        else:
            if p:
                tags.append(p)
    return tags, phantom


def parse_eligible_shifts(raw: str, cicu_shift: str | None) -> set | None:
    """Parse EligibleShifts cell → permitted shift set, or None (= no restriction).
    'CICU' maps to cicu_shift name.
    """
    if not raw or pd.isna(raw):
        return None
    allowed = set()
    for p in str(raw).upper().split(","):
        p = p.strip()
        if p == "CICU" and cicu_shift:
            allowed.add(cicu_shift)
        elif p:
            allowed.add(p)
    return allowed or None


# ─────────────────────────────────────────────────────────────────────────────
# BASE SCHEDULER
# ─────────────────────────────────────────────────────────────────────────────

class RosterScheduler:
    """Base class: all shared constraints. Subclasses override _apply_staff_rules_impl
    and _setup_fairness_impl."""

    def __init__(self, csv_path: str, output_dir: str, config: dict,
                 template_path: str = None, prev_month_path: str = None):
        self.csv_path        = csv_path
        self.output_dir      = output_dir
        self.cfg             = copy.deepcopy(config)
        self.template_path   = template_path
        self.prev_month_path = prev_month_path
        self.public_holidays: list[str] = []
        self.base_year: int | None = None
        self.prev_month_data:  dict[str, dict] = {}
        self._prev_month_rows: list = []
        self._prev_month_hdrs: list = []

        self.staff_data  = None
        self.dates: list[str] = []
        self.weekdays:  list[str] = []
        self.saturdays: list[str] = []
        self.sundays:   list[str] = []
        self.phs:       list[str] = []
        self.fridays:   set[str]  = set()
        self.pre_phs:   set[str]  = set()
        self._date_cache:  dict[str, datetime | None] = {}
        self._dtype_cache: dict[str, str] = {}
        self._date_idx:    dict[str, int] = {}
        self._all_shifts:  list[str] = []

    # ── Shift-category helpers ────────────────────────────────────────────────

    def _sc_names(self, *cats) -> list[str]:
        """Return all shift names for the given shift_categories."""
        sc = self.cfg["shift_categories"]
        return [n for c in cats for n in sc.get(c, {}).get("names", [])]

    @property
    def _cicu_name(self) -> str | None:
        names = self.cfg["shift_categories"].get("cicu", {}).get("names", [])
        return names[0] if names else None

    @property
    def _hybrid_shift_name(self) -> str | None:
        names = self.cfg["shift_categories"].get("hybrid_shift", {}).get("names", [])
        return names[0] if names else None

    @property
    def _call_count(self) -> list[str]:
        """All shifts that count as a call — everything except WR and SB."""
        return self._sc_names("main", "weekend_extra", "cicu", "hybrid_shift", "half")

    # ── Public API ────────────────────────────────────────────────────────────

    def set_public_holidays(self, holidays: list[str]):
        self.public_holidays = list(holidays)
        self._dtype_cache.clear()

    def set_base_year(self, year: int | None):
        """Anchor year for yearless date headers (e.g. '3-Jun')."""
        self.base_year = int(year) if year else None
        self._date_cache.clear()
        self._dtype_cache.clear()

    # ── Date helpers ──────────────────────────────────────────────────────────

    def _parse(self, dc: str) -> datetime | None:
        if dc not in self._date_cache:
            self._date_cache[dc] = _parse_date(dc)
        return self._date_cache[dc]

    @staticmethod
    def _raw_date_parts(dc: str) -> tuple[int, int, int | None] | None:
        """Parse a header into (day, month, year|None) without guessing a year."""
        dc = str(dc).strip()
        for sep in ("-", " "):
            parts = dc.split(sep)
            if len(parts) >= 2 and parts[0].isdigit() and parts[1] in _MONTH_MAP:
                year = None
                if len(parts) >= 3 and parts[2].isdigit():
                    y = int(parts[2]); year = y + 2000 if y < 100 else y
                return int(parts[0]), _MONTH_MAP[parts[1]], year
        if dc.count("/") == 2:
            d, m, y = dc.split("/")
            if d.isdigit() and m.isdigit() and y.isdigit():
                yi = int(y); yi += 2000 if yi < 100 else 0
                return int(d), int(m), yi
        r = pd.to_datetime(dc, errors="coerce")
        return None if pd.isna(r) else (r.day, r.month, r.year)

    def _resolve_date_headers(self, raw_cols: list[str]):
        """Pre-resolve date headers into self._date_cache using one anchor year,
        rolling forward on a Dec→Jan month wrap. Yearless headers fall back to
        base_year, then today's year."""
        parsed = [(dc, self._raw_date_parts(dc)) for dc in raw_cols]
        anchor_year = next((p[2] for _, p in parsed if p and p[2] is not None), None)
        if anchor_year is None:
            anchor_year = self.base_year or datetime.today().year
        anchor_month = next((p[1] for _, p in parsed if p), 1)
        for dc, p in parsed:
            if not p:
                self._date_cache[dc] = None
                continue
            day, month, yr = p
            if yr is None:
                yr = anchor_year + (1 if month < anchor_month else 0)
            try:
                self._date_cache[dc] = datetime(yr, month, day)
            except ValueError:
                self._date_cache[dc] = None

    def _is_ph(self, dc: str) -> bool:
        dt = self._parse(dc)
        return bool(dt and _ph_key(dt) in self.public_holidays)

    def _day_type(self, dc: str) -> str:
        """Returns: 'weekday'|'friday'|'saturday'|'sunday'|'ph'. Cached."""
        if dc not in self._dtype_cache:
            if self._is_ph(dc):
                t = "ph"
            else:
                dt = self._parse(dc)
                wd = dt.weekday() if dt else -1
                t = ("saturday" if wd == 5 else "sunday" if wd == 6
                     else "friday" if wd == 4 else "weekday")
            self._dtype_cache[dc] = t
        return self._dtype_cache[dc]

    def _is_wknd(self, dc: str) -> bool:
        return self._day_type(dc) in ("saturday","sunday","ph")

    def _day_name(self, dc: str) -> str:
        dt = self._parse(dc)
        return dt.strftime("%A").lower() if dt else ""

    def _prev_day_is_sunday(self, dc: str) -> bool:
        dt = self._parse(dc)
        return bool(dt and (dt - timedelta(days=1)).weekday() == 6)

    def _req_key(self, dc: str) -> str:
        """Normalise friday → weekday for slot lookups."""
        dt = self._day_type(dc)
        return "weekday" if dt == "friday" else dt

    # ── Data loading ──────────────────────────────────────────────────────────

    def _staff_sort_keys(self) -> list[str]:
        """Columns to sort staff by before solving. Override in subclasses."""
        return ["Team", "Ward", "Name"]

    def load_data(self):
        cfg = self.cfg
        self.staff_data = pd.read_csv(self.csv_path, header=1)
        self.staff_data.columns = [str(c).strip() for c in self.staff_data.columns]

        n_before = len(self.staff_data)
        self.staff_data = self.staff_data[
            self.staff_data["Name"].notna() & (self.staff_data["Name"].astype(str).str.strip() != "")
        ].reset_index(drop=True)
        if n_before != len(self.staff_data):
            print(f"Dropped {n_before - len(self.staff_data)} blank rows "
                  f"({len(self.staff_data)} staff remain)")

        missing = [c for c in cfg["required_cols"] if c not in self.staff_data.columns]
        if missing:
            raise ValueError(f"Missing CSV columns: {missing}")

        sort_keys = [k for k in self._staff_sort_keys() if k in self.staff_data.columns]
        if sort_keys:
            self.staff_data = self.staff_data.sort_values(sort_keys).reset_index(drop=True)

        start = cfg["date_col_start"]
        if len(self.staff_data.columns) <= start:
            raise ValueError("No date columns found")

        self.dates = []
        self.weekdays, self.saturdays, self.sundays = [], [], []
        self.phs, self.fridays, self.pre_phs = [], set(), set()

        self._resolve_date_headers(list(self.staff_data.columns)[start:])
        for dc in list(self.staff_data.columns)[start:]:
            if self._parse(dc) is None:
                print(f"Warning: cannot parse '{dc}' — skipped")
                continue
            self.dates.append(dc)
            dt = self._day_type(dc)
            if dt == "saturday":  self.saturdays.append(dc)
            elif dt == "sunday":  self.sundays.append(dc)
            elif dt == "ph":      self.phs.append(dc)
            else:                 self.weekdays.append(dc)
            if dt == "friday":    self.fridays.add(dc)

        self.pre_phs = {
            dc for i, dc in enumerate(self.dates[:-1])
            if dc in self.weekdays and self._is_ph(self.dates[i+1])
        }
        self._date_idx   = {dc: i for i, dc in enumerate(self.dates)}
        self._all_shifts = self._sc_names(
            "main", "weekend_extra", "cicu", "half", "hybrid_shift", "wr", "sb"
        )

        self._preflight_check()

        print(f"Loaded {len(self.staff_data)} staff, {len(self.dates)} days "
              f"({len(self.weekdays)}wd {len(self.saturdays)}sat "
              f"{len(self.sundays)}sun {len(self.phs)}ph)")

        if self.prev_month_path:
            self._read_prev_month()

    def _preflight_check(self):
        qs = self._cicu_name
        if not qs or "EligibleShifts" not in self.staff_data.columns:
            return
        eligible = sum(
            1 for _, row in self.staff_data.iterrows()
            if "CICU" in str(row.get("EligibleShifts","")).upper()
        )
        has_cicu  = bool(self.cfg["shift_categories"].get("cicu", {}).get("names"))
        days_need = len(self.dates) if has_cicu else 0
        if eligible == 0 and days_need > 0:
            print(f"WARNING: No CICU-eligible staff but {qs} required on "
                  f"{days_need} days → likely INFEASIBLE")

    # ── Call points ───────────────────────────────────────────────────────────

    def _points_for(self, dc: str, shift: str) -> int:
        """Scaled call points for (date, shift). WR/SB/half → 0 except ho6_points."""
        cfg = self.cfg
        sc  = cfg["shift_categories"]
        if (shift in sc.get("wr",  {}).get("names", []) or
                shift in sc.get("sb",  {}).get("names", [])):
            return 0
        dt = self._day_type(dc)
        if dt == "weekday" and dc in self.pre_phs and "pre_ph" in cfg["call_points"]:
            dt = "pre_ph"
        if shift == self._hybrid_shift_name and "r3_points" in cfg:
            return cfg["r3_points"].get(dt, cfg["r3_points"].get("weekday", 0))
        if shift in sc.get("half", {}).get("names", []):
            return cfg.get("ho6_points", 0)
        return cfg["call_points"].get(dt, cfg["call_points"].get("weekday", 0))

    # ── WR slots for a date ───────────────────────────────────────────────────

    def _wr_slots(self, dc: str) -> int:
        wr = self.cfg["shift_categories"].get("wr", {})
        rk = self._req_key(dc)
        if rk not in ("saturday", "sunday", "ph"):
            return 0
        base = wr.get("slots", {}).get(rk, 0)
        if rk == "ph" and self._prev_day_is_sunday(dc):
            return wr.get("ph_after_sunday", base)
        return base

    # ── Model building ────────────────────────────────────────────────────────

    def create_model(self):
        if self.staff_data is None:
            raise RuntimeError("Call load_data() before create_model()")

        model = cp_model.CpModel()
        N, D, SH = len(self.staff_data), self.dates, self._all_shifts
        print(f"\n=== Building Model ({N} staff, {len(D)} days, {len(SH)} shifts) ===")

        sv = {(s,d,sh): model.NewBoolVar(f"x{s}_{d}_{sh}")
              for s in range(N) for d in D for sh in SH}

        worked = [
            [sum(sv[(s,d,sh)] for sh in SH) for d in D]
            for s in range(N)
        ]

        self._enforce_max_one_shift_per_day(model, worked)
        self._add_daily_requirements(model, sv, N, D)
        self._add_rest_constraints(model, sv, worked, N, D)
        self._add_eligible_shift_gating(model, sv, N)
        self._add_hard_blocks(model, sv, N)
        penalties = self._add_soft_penalties(model, sv, N)
        self._apply_staff_rules_impl(model, sv, N)
        self._add_team_preferences(model, sv, N, penalties)
        fairness_terms = self._setup_fairness_impl(model, sv, N)

        model.Minimize(sum(p*w for p,w in penalties) + sum(fairness_terms))
        return model, sv

    def _enforce_max_one_shift_per_day(self, model, worked):
        for _, row in enumerate(worked):
            for w in row:
                model.Add(w <= 1)

    def _add_daily_requirements(self, model, sv, N, D):
        sc = self.cfg["shift_categories"]

        col_sum = {(d,sh): sum(sv[(s,d,sh)] for s in range(N))
                   for d in D for sh in self._all_shifts}

        for d in D:
            rk      = self._req_key(d)
            is_wknd = self._is_wknd(d)

            # Fill each main and r3 shift exactly once every day
            for cat in ("main", "hybrid_shift"):
                for sh in sc.get(cat, {}).get("names", []):
                    model.Add(col_sum[(d, sh)] == 1)

            # weekend_extra: exactly once on weekends/PH, absent on weekdays
            for sh in sc.get("weekend_extra", {}).get("names", []):
                model.Add(col_sum[(d, sh)] == (1 if is_wknd else 0))

            # cicu: exactly once every day (eligibility gated separately)
            for sh in sc.get("cicu", {}).get("names", []):
                model.Add(col_sum[(d, sh)] == 1)

            # WR: named slots filled in order up to wr_slots count
            wr_slots = self._wr_slots(d)
            for i, sh in enumerate(sc.get("wr", {}).get("names", [])):
                model.Add(col_sum[(d, sh)] == (1 if i < wr_slots else 0))

            # SB and half: total equals configured slot count
            for cat in ("sb", "half"):
                entry = sc.get(cat, {})
                names = entry.get("names", [])
                if names:
                    slots = entry.get("slots", {}).get(rk, 0)
                    model.Add(sum(col_sum[(d, sh)] for sh in names) == slots)

    def _add_rest_constraints(self, model, sv, worked, N, D):
        """1 rest day after WR; 2 rest days after all other shifts."""
        wr_set    = set(self._sc_names("wr"))
        non_wr_sh = [sh for sh in self._all_shifts if sh not in wr_set]

        for s in range(N):
            # Day after: blocked after any shift
            for i in range(len(D) - 1):
                model.Add(worked[s][i] + worked[s][i+1] <= 1)
            # Two days after: blocked only after a non-WR shift
            non_wr_worked = [sum(sv[(s,d,sh)] for sh in non_wr_sh) for d in D]
            for i in range(len(D) - 2):
                model.Add(non_wr_worked[i] + worked[s][i+2] <= 1)

    def _add_eligible_shift_gating(self, model, sv, N):
        """Apply EligibleShifts constraints per eligible_shifts_mode:
          "none"           — no gating
          "cicushift_only" — block cicu shift unless CICU in EligibleShifts
          "whitelist"      — block any shift not listed in EligibleShifts
        """
        mode = self.cfg.get("eligible_shifts_mode", "none")
        if mode == "none":
            return
        qs  = self._cicu_name
        col = "EligibleShifts" if "EligibleShifts" in self.staff_data.columns else None

        for s, staff in self.staff_data.iterrows():
            raw     = str(staff.get(col,"") if col else "").strip()
            allowed = parse_eligible_shifts(raw, qs)

            if mode == "cicushift_only":
                if qs and (allowed is None or qs not in allowed):
                    for d in self.dates:
                        model.Add(sv[(s,d,qs)] == 0)

            elif mode == "whitelist":
                if allowed is not None:
                    for sh in self._all_shifts:
                        if sh not in allowed:
                            for d in self.dates:
                                model.Add(sv[(s,d,sh)] == 0)
                elif qs:
                    for d in self.dates:
                        model.Add(sv[(s,d,qs)] == 0)

    def _add_hard_blocks(self, model, sv, N):
        cfg    = self.cfg
        blocks = set(cfg["hard_blocks"])
        leave  = set(cfg["leave_values"])

        for s, staff in self.staff_data.iterrows():
            for d in self.dates:
                val = str(staff.get(d,"")).strip().upper()
                if not val or val == "NAN":
                    continue
                if val in leave:
                    self._block_day(model, sv, s, d)
                    idx = self._date_idx[d]
                    if idx > 0:
                        self._block_day(model, sv, s, self.dates[idx-1])
                elif val in blocks:
                    self._block_day(model, sv, s, d)

    def _add_soft_penalties(self, model, sv, N) -> list:
        pen_cfg   = self.cfg["soft_penalties"]
        all_sh    = self._all_shifts
        main_sh   = self._call_count
        penalties = []

        for s, staff in self.staff_data.iterrows():
            for d in self.dates:
                val = str(staff.get(d,"")).strip().upper()
                if val not in pen_cfg or pen_cfg[val] == 0:
                    continue
                w = pen_cfg[val]
                if val == "REQUEST":
                    # fulfilled only by a main shift
                    p = model.NewBoolVar(f"pr_{s}_{d}")
                    model.Add(sum(sv[(s,d,sh)] for sh in main_sh) + p >= 1)
                    penalties.append((p, w))
                else:
                    # any shift triggers the penalty
                    penalties.append((sum(sv[(s,d,sh)] for sh in all_sh), w))

        print(f"Soft penalties: {len(penalties)} terms")
        return penalties

    def _add_team_preferences(self, model, sv, N, penalties: list):
        cfg   = self.cfg
        prefs = cfg.get("team_preferences", {})
        w     = cfg.get("team_pref_penalty", 2)
        if not prefs or "Team" not in self.staff_data.columns:
            return

        call_shifts    = self._call_count
        non_pref_by_team = {
            team: [sh for sh in call_shifts if sh not in pref]
            for team, pref in prefs.items()
        }

        for s, staff in self.staff_data.iterrows():
            team     = str(staff.get("Team","")).strip().upper()
            non_pref = non_pref_by_team.get(team, [])
            for sh in non_pref:
                for d in self.dates:
                    penalties.append((sv[(s,d,sh)], w))

    # ── Limits helpers ────────────────────────────────────────────────────────

    def _add_wr_limit(self, model, sv, s: int):
        mx    = self.cfg["limits"].get("wr_max", 2)
        wr_sh = self._sc_names("wr")
        model.Add(sum(sv[(s,d,sh)] for d in self.dates for sh in wr_sh) <= mx)

    def _add_sb_limit(self, model, sv, s: int):
        mx    = self.cfg["limits"].get("sb_max", 0)
        sb_sh = self._sc_names("sb")
        if mx and sb_sh:
            model.Add(sum(sv[(s,d,sh)] for d in self.dates for sh in sb_sh) <= mx)

    def _block_day(self, model, sv, s: int, d: str):
        for sh in self._all_shifts:
            model.Add(sv[(s, d, sh)] == 0)

    def _block_shifts(self, model, sv, s: int, d: str, shifts):
        for sh in shifts:
            model.Add(sv[(s, d, sh)] == 0)

    # ── Fairness helpers ──────────────────────────────────────────────────────

    def _minmax_range(self, model, var_list: list, upper: int, label: str):
        if not var_list:
            return 0
        mx = model.NewIntVar(0, upper, f"mx_{label}")
        mn = model.NewIntVar(0, upper, f"mn_{label}")
        model.AddMaxEquality(mx, var_list)
        model.AddMinEquality(mn, var_list)
        return mx - mn

    def _make_pts_var(self, model, sv, s: int, shifts: list,
                      label: str, upper: int = 600):
        terms = [sv[(s,d,sh)] * self._points_for(d,sh)
                 for d in self.dates for sh in shifts]
        v = model.NewIntVar(0, upper, f"pts_{label}")
        model.Add(v == sum(terms))
        return v

    def _make_count_var(self, model, sv, s: int, days: list, shifts: list,
                        upper: int, label: str):
        v = model.NewIntVar(0, upper, label)
        model.Add(v == sum(sv[(s,d,sh)] for d in days for sh in shifts))
        return v

    def _build_fairness_terms(self, model, vmap: dict, weights: dict, label: str) -> list:
        return [self._minmax_range(model, vl, up, f"{k}_{label}") * weights.get(k, 0)
                for k, (vl, up) in vmap.items() if weights.get(k, 0) and vl]

    # ── Subclass hooks ────────────────────────────────────────────────────────

    def _apply_staff_rules_impl(self, model, sv, N):
        raise NotImplementedError

    def _setup_fairness_impl(self, model, sv, N) -> list:
        raise NotImplementedError

    # ── Solve & export ────────────────────────────────────────────────────────

    def solve_and_export(self, time_limit: float = 300.0, workers: int = 8):
        if self.staff_data is None:
            raise RuntimeError("Call load_data() first")
        model, sv = self.create_model()

        print(f"\n=== Solving (max {time_limit:.0f}s, {workers} workers) ===", flush=True)
        class _Progress(cp_model.CpSolverSolutionCallback):
            def on_solution_callback(self):
                print(f"  [{self.WallTime():.1f}s] solution obj={self.ObjectiveValue():.0f}", flush=True)

        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds   = time_limit
        solver.parameters.num_search_workers    = workers
        solver.parameters.log_search_progress   = True
        solver.parameters.log_to_stdout         = False
        status = solver.Solve(model, _Progress())

        if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            print(f"No solution. Status: {solver.StatusName(status)}")
            return None

        print(f"{'OPTIMAL' if status==cp_model.OPTIMAL else 'FEASIBLE'} "
              f"in {solver.WallTime():.1f}s")

        df = self._build_output(solver, sv)
        self._check_violations(solver, sv)
        self._display_statistics(df, solver, sv)

        os.makedirs(self.output_dir, exist_ok=True)
        dt0  = self._parse(self.dates[0])
        stem = (f"{self.cfg['roster_type']}_{dt0.strftime('%b_%Y')}_roster"
                if dt0 else f"{self.cfg['roster_type']}_roster")

        csv_path = os.path.join(self.output_dir, stem + ".csv")
        df.to_csv(csv_path, index=False)
        print(f"CSV: {csv_path}")

        if self.template_path:
            xl_path = os.path.join(self.output_dir, stem + ".xlsx")
            self._export_excel(df, solver, sv, xl_path)
            print(f"Excel: {xl_path}")

        return df

    def _get_assigned(self, solver, sv, s: int, d: str) -> str:
        return next((sh for sh in self._all_shifts if solver.Value(sv[(s, d, sh)])), "")

    def _build_output(self, solver, sv) -> pd.DataFrame:
        rows = []
        for s, staff in self.staff_data.iterrows():
            row = {c: staff[c] for c in self.cfg["required_cols"]
                   if c in self.staff_data.columns}
            for d in self.dates:
                row[d] = self._get_assigned(solver, sv, s, d)
            rows.append(row)
        return pd.DataFrame(rows)

    def _check_violations(self, solver, sv):
        print("\n=== Hard Violations ===")
        leave  = set(self.cfg["leave_values"])
        blocks = set(self.cfg["hard_blocks"])
        found  = False
        for s, staff in self.staff_data.iterrows():
            name = staff.get("Name", f"Staff{s}")
            for d in self.dates:
                val      = str(staff.get(d,"")).strip().upper()
                assigned = self._get_assigned(solver, sv, s, d)
                if assigned and (val in leave or val in blocks):
                    print(f"  HARD: {name} → {assigned} on {d} ({val})")
                    found = True
        if not found:
            print("  None")

    def _display_statistics(self, df, solver, sv):
        pass

    @staticmethod
    def _rng(lst) -> str:
        return f"{min(lst):.1f}–{max(lst):.1f}  (range {max(lst)-min(lst):.1f})"

    def _get_tags(self, staff) -> tuple[list, int]:
        return parse_phantom_points(staff.get("SpecialReq", ""))

    def _print_soft_violations(self, solver, sv):
        pen_cfg    = self.cfg["soft_penalties"]
        violations = []
        for s, staff in self.staff_data.iterrows():
            name = staff.get("Name", f"Staff{s}")
            for d in self.dates:
                val = str(staff.get(d, "")).strip().upper()
                if val not in pen_cfg:
                    continue
                assigned = self._get_assigned(solver, sv, s, d)
                if val == "REQUEST":
                    if not assigned:
                        violations.append(f"  REQUEST not fulfilled: {name} on {d}")
                elif assigned:
                    violations.append(f"  {val}: {name} on {d} → {assigned}")
        print("\nSoft Constraint Violations:")
        for v in violations:
            print(v)
        if not violations:
            print("  None")

    def _export_excel(self, df, solver, sv, xl_path: str):
        colors = {
            "LEAVE":     "D0CECE",
            "AUTOBLOCK": "D0CECE",
            "POSTCALL":"FFF2CC",
            "COURSE":    "E1BEE7",
            "SUBSPEC":   "C9D7F8",
            "REQUEST":   "E2EFDA",
            "BLOCK":     "FFCDD2",
            "NEW":       "FFF9C4",
            "weekend":   "E7E6E6",
        }
        prio = self.cfg["color_priority"]

        shutil.copy2(self.template_path, xl_path)
        wb = load_workbook(xl_path)
        if self.prev_month_path:
            self._populate_prev_month(wb)

        skip = {"Lookup Data","Index","Month-1"}
        sheets = [n for n in wb.sheetnames if n not in skip]
        if not sheets:
            print("ERROR: No roster sheet in template"); return
        ws = wb[sheets[0]]

        dt0 = self._parse(self.dates[0])
        if dt0:
            ws["B1"] = dt0.strftime("%B %Y")
            ws["I5"] = dt0

        cap = self._template_date_capacity(ws)
        if len(self.dates) > cap:
            print(f"WARNING: {len(self.dates)} dates exceed template capacity "
                  f"({cap}); extra days not written to Excel")

        for col_idx, d in enumerate(self.dates):
            if col_idx >= cap:
                break
            col_letter = get_column_letter(9 + col_idx)
            if self._is_ph(d):
                ws[f"{col_letter}1"] = "Y"
            elif d in self.pre_phs:
                ws[f"{col_letter}1"] = "PRE"

        skip_cols = set(self.cfg.get("export_skip_cols", []))
        meta_cols = [c for c in self.cfg["required_cols"] if c not in skip_cols]
        for row_idx, (s, staff) in enumerate(self.staff_data.iterrows()):
            row = 6 + row_idx
            for ci, col in enumerate(meta_cols):
                ws.cell(row=row, column=2+ci, value=staff.get(col,""))
            for col_idx, d in enumerate(self.dates):
                if col_idx >= cap:
                    break
                col  = 9 + col_idx
                cell = ws.cell(row=row, column=col)
                val  = str(staff.get(d,"")).strip().upper()
                cell.value = self._get_assigned(solver, sv, s, d)
                hex_col = next((colors[k] for k in prio if k in val), None)
                if not hex_col and self._is_wknd(d):
                    hex_col = colors.get("weekend")
                if hex_col:
                    cell.fill = PatternFill(start_color=hex_col, end_color=hex_col,
                                            fill_type="solid")
        wb.save(xl_path)

    def _template_date_capacity(self, ws, default: int = 36) -> int:
        """Date columns available before summary columns (scans row 5 from col I)."""
        markers = {"total # of calls", "total call points", "total months"}
        for col in range(9, 9 + 200):
            v = str(ws.cell(row=5, column=col).value or "").strip().lower()
            if v in markers:
                return col - 9
        return default

    def _read_prev_month(self):
        if not self.prev_month_path or not os.path.exists(self.prev_month_path):
            return
        self.prev_month_data  = {}
        self._prev_month_rows = []
        try:
            prev_wb = load_workbook(self.prev_month_path, data_only=True)
            skip    = {"Lookup Data", "Index", "Month-1"}
            sheets  = [n for n in prev_wb.sheetnames if n not in skip]
            if not sheets:
                prev_wb.close()
                return
            prev_ws = prev_wb[sheets[0]]

            date_cols, calls_col, cum_pts_col, cum_months_col = [], None, None, None
            for col in range(7, 200):
                v = prev_ws.cell(row=5, column=col).value
                vs = str(v).strip().lower()
                if _parse_date(str(v)):
                    date_cols.append(col)
                elif vs == "total # of calls":
                    calls_col = col
                elif vs == "total call points":
                    cum_pts_col = col
                elif vs == "total months":
                    cum_months_col = col

            last2       = date_cols[-2:] if len(date_cols) >= 2 else date_cols
            self._prev_month_hdrs = [prev_ws.cell(row=5, column=lc).value for lc in last2]
            print(f"  Month-1: last date cols {last2}, calls col {calls_col}, "
                   f"cum_pts col {cum_pts_col}, "
                  f"months col {cum_months_col}")

            # Detect staff end: two rows above where col E equals first main shift
            first_shift = self._sc_names("main")[0].upper()
            staff_end   = None
            for r in range(7, 200):
                if str(prev_ws.cell(row=r, column=5).value or "").strip().upper() == first_shift:
                    staff_end = r - 2
                    break
            if staff_end is None:
                print("  Warning: staff end row not detected — reading until blank name")
            print(f"  Staff rows: 6 to {staff_end or 'end'}")

            src_end = (staff_end + 1) if staff_end else 200
            for src_row in range(6, src_end):
                name = prev_ws.cell(row=src_row, column=2).value
                if not name or not str(name).strip():
                    break
                name       = str(name).strip()
                last2_vals = [prev_ws.cell(row=src_row, column=lc).value for lc in last2]
                calls_val  = (prev_ws.cell(row=src_row, column=calls_col).value or 0) if calls_col else 0
                points_val = (prev_ws.cell(row=src_row, column=cum_pts_col).value or 0) if cum_pts_col else 0
                self._prev_month_rows.append((name, last2_vals, calls_val, points_val))

                cum_pts    = prev_ws.cell(row=src_row, column=cum_pts_col).value if cum_pts_col else None
                cum_months = prev_ws.cell(row=src_row, column=cum_months_col).value if cum_months_col else None
                if cum_pts is not None or cum_months is not None:
                    self.prev_month_data[name] = {
                        "cum_points": float(cum_pts or 0),
                        "months":     int(cum_months or 0),
                    }

            prev_wb.close()
            print(f"  Month-1: {len(self._prev_month_rows)} staff read, "
                  f"{len(self.prev_month_data)} with cumulative data")
        except Exception as e:
            print(f"Warning: prev month read failed: {e}")

    def _populate_prev_month(self, wb):
        if not self._prev_month_rows:
            return
        try:
            if "Month-1" not in wb.sheetnames:
                wb.create_sheet("Month-1")
            m1 = wb["Month-1"]

            for ci, hdr in enumerate(self._prev_month_hdrs):
                m1.cell(row=2, column=3 + ci, value=hdr)

            for r in range(3, m1.max_row + 1):
                for c in range(2, 8):
                    m1.cell(row=r, column=c, value=None)

            for out_row, (name, last2_vals, calls_val, points_val) in \
                    enumerate(self._prev_month_rows, start=3):
                m1.cell(row=out_row, column=2, value=name)
                for ci, val in enumerate(last2_vals):
                    m1.cell(row=out_row, column=3 + ci, value=val)
                m1.cell(row=out_row, column=5, value=calls_val)
                m1.cell(row=out_row, column=6, value=points_val)
                months = self.prev_month_data.get(name, {}).get("months", 0)
                m1.cell(row=out_row, column=7, value=months)

            print(f"  Month-1: {len(self._prev_month_rows)} staff written")
        except Exception as e:
            print(f"Warning: Month-1 write failed: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# MO SUBCLASS
# ─────────────────────────────────────────────────────────────────────────────

class MORosterScheduler(RosterScheduler):

    def _get_tags(self, staff) -> tuple[list, int]:
        tags, phantom = parse_phantom_points(staff.get("SpecialReq",""))
        if "NEW" in tags:
            phantom += self.cfg.get("new_phantom_points", 4)
        return tags, phantom

    def _apply_staff_rules_impl(self, model, sv, N):
        cfg    = self.cfg
        mo_sh  = self._call_count
        wr_sh  = self._sc_names("wr")
        sb_sh  = self._sc_names("sb")
        all_sh = self._all_shifts
        wknd   = self.saturdays + self.sundays + self.phs

        for s, staff in self.staff_data.iterrows():
            subspec = str(staff.get("Subspec","")).strip().upper()
            tags, _ = self._get_tags(staff)
            self._add_wr_limit(model, sv, s)
            self._add_sb_limit(model, sv, s)

            if subspec == "FAMMED":
                for d in self.dates:
                    self._block_shifts(model, sv, s, d, mo_sh + sb_sh)
                wr_tot = sum(sv[(s,d,sh)] for d in self.dates for sh in wr_sh)
                model.Add(wr_tot >= 1); model.Add(wr_tot <= 2)

            elif subspec == "ELECTIVES":
                for d in self.dates:
                    self._block_shifts(model, sv, s, d, wr_sh + sb_sh)
                model.Add(sum(sv[(s,d,sh)] for d in self.dates for sh in mo_sh) == 2)
                sat_mo = [sv[(s,d,sh)] for d in self.saturdays for sh in mo_sh]
                fri_pp = [sv[(s,d,sh)] for d in self.dates
                          if d in self.fridays or d in self.pre_phs for sh in mo_sh]
                if sat_mo: model.Add(sum(sat_mo) == 1)
                if fri_pp: model.Add(sum(fri_pp) == 1)

            if subspec in ("DCD","AMBU"):
                sun_thu = [d for d in self.dates if self._day_name(d) in
                           ("sunday","monday","tuesday","wednesday","thursday")]
                mo_st = [sv[(s,d,sh)] for d in sun_thu for sh in mo_sh]
                if mo_st: model.Add(sum(mo_st) <= 2)

            if "HC" in tags:
                for d in self.dates:
                    self._block_shifts(model, sv, s, d, wr_sh + sb_sh)
                wknd_mo = [sv[(s,d,sh)] for d in wknd for sh in mo_sh]
                wkdy_mo = [sv[(s,d,sh)] for d in self.weekdays for sh in mo_sh]
                if wknd_mo: model.Add(sum(wknd_mo) == 1)
                if wkdy_mo: model.Add(sum(wkdy_mo) == 2)

            if "NEW" in tags:
                n = cfg.get("new_block_days", 10)
                for d in self.dates[:min(n, len(self.dates))]:
                    self._block_day(model, sv, s, d)

            if "VIKAS" in tags:
                qs     = self._cicu_name
                non_qs = [sh for sh in all_sh if sh != qs]
                for d in self.dates:
                    self._block_shifts(model, sv, s, d, non_qs)
                    if self._day_name(d) not in cfg.get("vikas_days", []):
                        model.Add(sv[(s, d, qs)] == 0)

        print("MO staff rules applied")

    def _setup_fairness_impl(self, model, sv, N) -> list:
        pool   = self.cfg["fairness_pools"][0]
        w      = pool["metrics"]
        mo_sh  = self._call_count
        wr_sh  = self._sc_names("wr")
        sb_sh  = self._sc_names("sb")
        excl_t = pool["exclude_subtypes"]
        excl_g = pool["exclude_tags"]

        pts_v, mo_v, sat_v, wr_v, sb_v = [], [], [], [], []

        for s, staff in self.staff_data.iterrows():
            subspec = str(staff.get("Subspec","")).strip().upper()
            tags, phantom = self._get_tags(staff)
            if subspec in excl_t or any(t in tags for t in excl_g):
                continue

            raw  = self._make_pts_var(model, sv, s, mo_sh, f"mo{s}")
            fair = model.NewIntVar(0, 600, f"fpts_mo{s}")
            model.Add(fair == raw + phantom)
            pts_v.append(fair)
            mo_v.append(self._make_count_var(model, sv, s, self.dates,     mo_sh, 20, f"mocnt{s}"))
            sat_v.append(self._make_count_var(model, sv, s, self.saturdays, mo_sh, 10, f"satcnt{s}"))
            wr_v.append(self._make_count_var(model, sv, s, self.dates,     wr_sh,  5, f"wrcnt_mo{s}"))
            sb_v.append(self._make_count_var(model, sv, s, self.dates,     sb_sh,  5, f"sbcnt_mo{s}"))

        vmap  = {"pts":(pts_v,600),"full_calls":(mo_v,20),"sat_calls":(sat_v,10),
                 "wr_count":(wr_v,5),"sb_count":(sb_v,5)}
        terms = self._build_fairness_terms(model, vmap, w, "mo")

        print(f"MO fairness: {len(pts_v)} regular staff, {len(terms)} active metrics")
        return terms

    def _display_statistics(self, _df, solver, sv):
        print("\n=== Statistics ===")
        mo_sh  = self._call_count
        wr_sh  = self._sc_names("wr")
        sb_sh  = self._sc_names("sb")
        pool   = self.cfg["fairness_pools"][0]
        excl_t = set(pool["exclude_subtypes"])
        excl_g = set(pool["exclude_tags"])
        scale  = self.cfg.get("call_points_scale", 1)

        pts_v, mo_v, sat_v, wr_v, sb_v = [], [], [], [], []
        special_n = 0

        for s, staff in self.staff_data.iterrows():
            subspec = str(staff.get("Subspec", "")).strip().upper()
            tags, _ = self._get_tags(staff)
            if subspec in excl_t or any(t in tags for t in excl_g):
                special_n += 1
                continue
            pts_v.append(sum(solver.Value(sv[(s, d, sh)]) * self._points_for(d, sh)
                             for d in self.dates for sh in mo_sh) / scale)
            mo_v.append(sum(solver.Value(sv[(s, d, sh)])
                            for d in self.dates for sh in mo_sh))
            sat_v.append(sum(solver.Value(sv[(s, d, sh)])
                             for d in self.saturdays for sh in mo_sh))
            wr_v.append(sum(solver.Value(sv[(s, d, sh)])
                            for d in self.dates for sh in wr_sh))
            sb_v.append(sum(solver.Value(sv[(s, d, sh)])
                            for d in self.dates for sh in sb_sh))

        print(f"\nRegular staff: {len(pts_v)}   Special: {special_n}")
        if pts_v:
            print(f"\nFairness (regular staff):")
            print(f"  Call points:    {self._rng(pts_v)}")
            print(f"  Full calls:     {self._rng(mo_v)}")
            print(f"  Saturday calls: {self._rng(sat_v)}")
            print(f"  WR shifts:      {self._rng(wr_v)}")
            if any(sb_v):
                print(f"  SB shifts:      {self._rng(sb_v)}")
        self._print_soft_violations(solver, sv)


# ─────────────────────────────────────────────────────────────────────────────
# HO SUBCLASS
# ─────────────────────────────────────────────────────────────────────────────

class HORosterScheduler(RosterScheduler):

    def _apply_staff_rules_impl(self, model, sv, N):
        ho_sh = self._sc_names("main", "half")
        wr_sh = self._sc_names("wr")

        for s, staff in self.staff_data.iterrows():
            tags, _ = self._get_tags(staff)
            self._add_wr_limit(model, sv, s)
            self._add_sb_limit(model, sv, s)

            if "NEW" in tags:
                # New HOs must cover ≥2 weekend slots in first 2 weeks
                first_2w = self.dates[:min(14, len(self.dates))]
                wknd_sh  = [sv[(s,d,sh)] for d in first_2w
                            if self._is_wknd(d) for sh in ho_sh + wr_sh]
                if wknd_sh:
                    model.Add(sum(wknd_sh) >= 2)

        print("HO staff rules applied")

    def _setup_fairness_impl(self, model, sv, N) -> list:
        cfg   = self.cfg
        pool  = cfg["fairness_pools"][0]
        w     = pool["metrics"]
        ho_sh = self._sc_names("main")
        ho6   = self._sc_names("half")
        wr_sh = self._sc_names("wr")
        sb_sh = self._sc_names("sb")
        scale = cfg.get("call_points_scale", 2)
        all_sh = self._all_shifts

        gw_triplets = [
            (i, i+1, i+2) for i in range(len(self.dates)-2)
            if all(self._parse(self.dates[j]) for j in (i,i+1,i+2)) and
               self._parse(self.dates[i]).weekday()   == 4 and
               self._parse(self.dates[i+1]).weekday() == 5 and
               self._parse(self.dates[i+2]).weekday() == 6
        ]
        print(f"HO golden weekend triplets: {len(gw_triplets)}")

        cross_month_w = w.get("cross_month", 0)
        global_avg    = 0.0
        if cross_month_w and self.prev_month_data:
            total_cum    = sum(v["cum_points"] for v in self.prev_month_data.values())
            total_months = sum(v["months"]     for v in self.prev_month_data.values())
            global_avg   = total_cum / total_months if total_months > 0 else 0.0
            print(f"HO cross-month: global avg {global_avg:.2f} pts/month "
                  f"({len(self.prev_month_data)} staff records)")

        pts_v, wr_v, sb_v, fc_v, gw_v, gap_v, dev_v = [], [], [], [], [], [], []

        for s, staff in self.staff_data.iterrows():
            _, phantom = self._get_tags(staff)
            name = str(staff.get("Name", "")).strip()

            pts_terms = ([sv[(s,d,sh)] * self._points_for(d,sh)
                          for d in self.dates for sh in ho_sh] +
                         [sv[(s,d,sh)] * cfg.get("ho6_points",2)
                          for d in self.dates for sh in ho6])
            raw  = model.NewIntVar(0, 1000, f"rawpts_ho{s}")
            fair = model.NewIntVar(0, 1000, f"fpts_ho{s}")
            model.Add(raw == sum(pts_terms))
            model.Add(fair == raw + phantom * scale)
            pts_v.append(fair)

            wr_v.append(self._make_count_var(model, sv, s, self.dates, wr_sh, 5, f"wrcnt_ho{s}"))
            sb_v.append(self._make_count_var(model, sv, s, self.dates, sb_sh, 5, f"sbcnt_ho{s}"))
            fc_v.append(self._make_count_var(model, sv, s, self.dates,
                self._call_count, 20, f"fccnt_ho{s}"))

            gw = model.NewIntVar(0, len(gw_triplets)+1, f"gw{s}")
            gw_bools = []
            for fi, si, ui in gw_triplets:
                b   = model.NewBoolVar(f"gwb{s}_{fi}")
                tot = sum(sv[(s,self.dates[j],sh)] for j in (fi,si,ui) for sh in all_sh)
                model.Add(tot == 0).OnlyEnforceIf(b)
                model.Add(tot >= 1).OnlyEnforceIf(b.Not())
                gw_bools.append(b)
            model.Add(gw == sum(gw_bools))
            gw_v.append(gw)

            day_worked = [sum(sv[(s,d,sh)] for sh in all_sh) for d in self.dates]
            excess = []
            for i in range(len(self.dates)-4):
                ex = model.NewIntVar(0, 5, f"ex{s}_{i}")
                model.Add(sum(day_worked[i:i+5]) - 2 <= ex)
                excess.append(ex)
            tot_ex = model.NewIntVar(0, 500, f"totex{s}")
            model.Add(tot_ex == sum(excess))
            gap_v.append(tot_ex)

            if cross_month_w and global_avg > 0:
                prev = self.prev_month_data.get(name)
                if prev and prev["months"] > 0:
                    target_scaled = max(0, int(
                        (global_avg * (prev["months"] + 1) - prev["cum_points"]) * scale
                    ))
                    dev = model.NewIntVar(0, 1000, f"cmdev_ho{s}")
                    model.Add(dev >= raw - target_scaled)
                    model.Add(dev >= target_scaled - raw)
                    dev_v.append(dev)

        vmap  = {"pts":(pts_v,1000),"full_calls":(fc_v,20),"wr_count":(wr_v,5),"sb_count":(sb_v,5),
                 "golden_wknds":(gw_v,len(gw_triplets)+1),"shift_spacing":(gap_v,500)}
        terms = self._build_fairness_terms(model, vmap, w, "ho")

        if dev_v:
            terms.append(sum(dev_v) * cross_month_w)
            print(f"  Cross-month: {len(dev_v)} staff with target deviations")

        print(f"HO fairness: {len(pts_v)} staff, {len(terms)} active metrics")
        return terms

    def _display_statistics(self, _df, solver, sv):
        print("\n=== Statistics ===")
        ho_sh = self._sc_names("main", "half")
        wr_sh = self._sc_names("wr")
        sb_sh = self._sc_names("sb")
        scale = self.cfg.get("call_points_scale", 2)

        pts_v, wr_v, sb_v = [], [], []
        for s, _ in self.staff_data.iterrows():
            pts_v.append(sum(solver.Value(sv[(s, d, sh)]) * self._points_for(d, sh)
                             for d in self.dates for sh in ho_sh) / scale)
            wr_v.append(sum(solver.Value(sv[(s, d, sh)])
                            for d in self.dates for sh in wr_sh))
            sb_v.append(sum(solver.Value(sv[(s, d, sh)])
                            for d in self.dates for sh in sb_sh))

        print(f"\nStaff: {len(pts_v)}")
        if pts_v:
            print(f"\nFairness:")
            print(f"  Call points: {self._rng(pts_v)}")
            print(f"  WR shifts:   {self._rng(wr_v)}")
            if any(sb_v):
                print(f"  SB shifts:   {self._rng(sb_v)}")
        self._print_soft_violations(solver, sv)


# ─────────────────────────────────────────────────────────────────────────────
# REG SUBCLASS
# ─────────────────────────────────────────────────────────────────────────────

class REGRosterScheduler(RosterScheduler):

    def _staff_sort_keys(self) -> list[str]:
        return ["StaffType", "Name"]

    def _stype(self, staff) -> str:
        return str(staff.get("StaffType","")).strip().upper()

    def _in_pool(self, staff, pool) -> bool:
        stype  = self._stype(staff)
        inc    = pool.get("include_types", [])
        excl_t = pool.get("exclude_subtypes", [])
        if stype == "AC": return False
        if inc and stype not in inc: return False
        if stype in excl_t: return False
        return True

    def _apply_staff_rules_impl(self, model, sv, N):
        all_sh = self._all_shifts
        qs     = self._cicu_name
        r3     = self._hybrid_shift_name
        wr_sh  = self._sc_names("wr")
        lims   = self.cfg["limits"]
        wknd   = self.saturdays + self.sundays + self.phs

        for s, staff in self.staff_data.iterrows():
            stype   = self._stype(staff)
            tags, _ = self._get_tags(staff)
            self._add_wr_limit(model, sv, s)

            if "CICU" in tags and qs:
                non_qs = [sh for sh in all_sh if sh != qs]
                for d in self.dates:
                    self._block_shifts(model, sv, s, d, non_qs)
                continue

            if stype == "AC":
                for d in wknd:
                    self._block_day(model, sv, s, d)
                for d in self.dates:
                    self._block_shifts(model, sv, s, d, wr_sh)
                total = sum(sv[(s,d,sh)] for d in self.dates for sh in all_sh)
                model.Add(total >= lims.get("ac_call_min",1))
                model.Add(total <= lims.get("ac_call_max",2))

            elif stype == "RP":
                if r3:
                    non_r3 = [sh for sh in all_sh if sh != r3]
                    for d in self.weekdays:
                        self._block_shifts(model, sv, s, d, non_r3)
                    model.Add(sum(sv[(s,d,r3)] for d in self.dates)
                              <= lims.get("rp_r3_max",3))
                non_wr = [sh for sh in all_sh if sh not in wr_sh]
                for d in wknd:
                    self._block_shifts(model, sv, s, d, non_wr)

        print("REG staff rules applied")

    def _setup_fairness_impl(self, model, sv, N) -> list:
        wr_sh   = self._sc_names("wr")
        pts_sh  = self._call_count
        terms   = []

        for pi, pool in enumerate(self.cfg["fairness_pools"]):
            w   = pool["metrics"]
            lbl = pool["label"].replace(" ","_")

            pts_v, sat_v, wr_v = [], [], []

            for s, staff in self.staff_data.iterrows():
                if not self._in_pool(staff, pool): continue

                tags, phantom = self._get_tags(staff)
                raw  = self._make_pts_var(model, sv, s, pts_sh, f"reg{pi}_{s}")
                fair = model.NewIntVar(0, 600, f"fpts_reg{pi}_{s}")
                model.Add(fair == raw + phantom)
                pts_v.append(fair)
                sat_v.append(self._make_count_var(
                    model, sv, s, self.saturdays, pts_sh, 10, f"satcnt_reg{pi}_{s}"))
                wr_v.append(self._make_count_var(
                    model, sv, s, self.dates, wr_sh, 5, f"wrcnt_reg{pi}_{s}"))

            vmap       = {"pts":(pts_v,600),"sat_calls":(sat_v,10),"wr_count":(wr_v,5)}
            pool_terms = self._build_fairness_terms(model, vmap, w, lbl)
            terms.extend(pool_terms)
            print(f"REG {pool['label']}: {len(pts_v)} staff, "
                  f"{len(pool_terms)} active metrics")

        return terms

    def _display_statistics(self, _df, solver, sv):
        print("\n=== Statistics ===")
        wr_sh  = self._sc_names("wr")
        pts_sh = self._call_count
        scale  = self.cfg.get("call_points_scale", 2)

        for pool in self.cfg["fairness_pools"]:
            print(f"\n[{pool['label']}]")
            pts_v, sat_v, wr_v = [], [], []
            for s, staff in self.staff_data.iterrows():
                if not self._in_pool(staff, pool): continue
                pts_v.append(sum(solver.Value(sv[(s, d, sh)]) * self._points_for(d, sh)
                                 for d in self.dates for sh in pts_sh) / scale)
                sat_v.append(sum(solver.Value(sv[(s, d, sh)])
                                 for d in self.saturdays for sh in pts_sh))
                wr_v.append(sum(solver.Value(sv[(s, d, sh)])
                                for d in self.dates for sh in wr_sh))
            print(f"  Staff: {len(pts_v)}")
            if pts_v:
                print(f"  Call points:    {self._rng(pts_v)}")
                print(f"  Saturday calls: {self._rng(sat_v)}")
                print(f"  WR shifts:      {self._rng(wr_v)}")

        self._print_soft_violations(solver, sv)


# ─────────────────────────────────────────────────────────────────────────────
# FACTORY
# ─────────────────────────────────────────────────────────────────────────────

_CLASSES = {"MO": MORosterScheduler, "HO": HORosterScheduler, "REG": REGRosterScheduler}

def build_scheduler(roster_type: str, csv_path: str, output_dir: str,
                    config: dict = None, template_path: str = None,
                    prev_month_path: str = None) -> RosterScheduler:
    """Factory — returns correct scheduler subclass.
    config defaults to a deep copy of ROSTER_CONFIGS[roster_type].
    """
    if roster_type not in _CLASSES:
        raise ValueError(f"roster_type must be one of {list(_CLASSES)}")
    cfg = config if config is not None else copy.deepcopy(ROSTER_CONFIGS[roster_type])
    return _CLASSES[roster_type](csv_path, output_dir, cfg, template_path, prev_month_path)


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    def _ask(prompt, default=""):
        v = input(f"{prompt} [{default}]: ").strip()
        return v or default

    print("=== Unified Roster Scheduler ===\n")
    rt      = _ask("Roster type (MO/HO/REG)").upper()
    csv_p   = _ask("Input CSV")
    out_dir = _ask("Output directory")
    tpl     = _ask("Excel template")
    prev    = _ask("Previous month xlsx (Enter to skip)", "") if tpl else ""

    year = int(_ask("Year", "e.g. 2026"))
    try:
        phs = get_singapore_ph(year)
    except ValueError as e:
        print(f"  {e}\n  Trying data.gov.sg…")
        phs = fetch_singapore_ph(year)
        if not phs:
            manual = _ask("Enter PHs manually (D-Mon-YY comma-sep, Enter for none)", "")
            phs = [h.strip() for h in manual.split(",") if h.strip()]
            if not phs:
                print("  WARNING: proceeding with NO public holidays")
    print(f"Public holidays {year}: {', '.join(phs)}")
    if _ask("Refresh from data.gov.sg? (y/n)").lower() == "y":
        fetched = fetch_singapore_ph(year)
        phs = fetched if fetched else phs
        print(f"  {'Refreshed' if fetched else 'Refresh failed — kept hardcoded'}: {', '.join(phs)}")
    ov = _ask("Add extra PHs (D-Mon-YY comma-sep, Enter to keep)", "")
    if ov and ov.lower() not in ("n", "no", "none"):
        phs += [h.strip() for h in ov.split(",")]

    cfg = copy.deepcopy(ROSTER_CONFIGS[rt])
    if _ask("Keep default fairness weights? (y/n)", "y").lower() != "y":
        print("\nFairness weights (Enter to keep):")
        for pool in cfg["fairness_pools"]:
            print(f"  [{pool['label']}]")
            for k, default in pool["metrics"].items():
                raw = input(f"    {k} [{default}]: ").strip()
                if raw.lstrip("-").isdigit():
                    pool["metrics"][k] = int(raw)

    if rt == "REG" and _ask("Enable R3? (y/n)", "y").lower() != "y":
        cfg["shift_categories"]["hybrid_shift"]["names"] = []

    s = build_scheduler(rt, csv_p, out_dir, cfg, tpl or None, prev or None)
    s.set_public_holidays(phs)
    s.set_base_year(year)
    s.load_data()
    s.solve_and_export()
